import openpyxl
# excel
# pipenv install openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
# wb.create_sheet("Sheet2", 0)    #0 before Sheet1
# wb.remove_sheet(sheet)
cell = sheet["a1"] # A1
print(f"{cell.value=}")
# cell.value = "asfdsdg"
print(f"{cell.row=}")
print(f"{cell.column=}")
print(f"{cell.coordinate=}")

cell = sheet.cell(row=1, column=1)
print(f"{sheet.max_row=}")
print(f"{sheet.max_column=}")
string: str = ""
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        string += str(cell.value) + "\t"
    string += "\n"
print(string)

column_a = sheet["a"]
print(f"{column_a=}")   # 4 cell objects tuple

cells = sheet["a:c"]
print(f"{cells=}")  # tuple in tuple

cells = sheet["a1:c3"]
cells = sheet["1"]  # 1st row
cells = sheet["1:3"]  # 1st-3rd row

sheet.append([1, 2, 3])
# sheet.insert_rows()
# sheet.insert_columns()
# sheet.delete_rows()
# sheet.delete_columns()
wb.save("1transactions2.xls")
print("row added")

# Command Query separation principle
